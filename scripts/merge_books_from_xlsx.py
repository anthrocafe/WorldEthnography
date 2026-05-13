#!/usr/bin/env python3
"""
Merge 总表 rows from 民族志信息表-*.xlsx into books-data.js:
- Match by (normalized author, year, English subtitle).
- Update title, author, year, publisher, summary, sourceField, locationEn,
  and when the book has no `sites`, also location + countryOrRegion.
- Never modify lat, lon, or sites on existing books.
- Append rows that have no match as new books (new ids + new coordinates).
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BOOKS_PATH = ROOT / "books-data.js"
PREFIX = "window.ETHNOGRAPHY_BOOKS = "


def norm_title_cell(t: str) -> str:
    t = str(t or "").replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", " ", t.strip())


def parse_cn_en_title(raw_title):
    text = norm_title_cell(raw_title)
    m = re.match(r"^《\s*(.*?)\s*》\s*\(\s*(.*?)\s*\)\s*$", text)
    if not m:
        return None, None
    return m.group(1).strip(), m.group(2).strip()


def canonical_title(cn: str, en: str) -> str:
    return f"《{cn}》({en})"


def norm_author(raw) -> str:
    a = str(raw or "").strip()
    a = a.replace("（", "(").replace("）", ")")
    a = re.sub(r"\([^)]*\)\s*$", "", a).strip()
    return re.sub(r"\s+", " ", a)


def match_key_from_book(book: dict):
    cn, en = parse_cn_en_title(book.get("title", ""))
    if not en:
        en = norm_title_cell(book.get("title", ""))
    return norm_author(book.get("author")), int(book.get("year", 0)), en


def match_key_from_row(title, author, year):
    cn, en = parse_cn_en_title(title)
    if not en:
        en = norm_title_cell(title)
    return norm_author(author), int(year), en


def parse_country_location(raw) -> tuple[str, str]:
    """Return (zh_line, en_line) for sourceField and locationEn derivation."""
    s = (raw or "")
    s = s.replace("\r\n", "\n").strip()
    if not s:
        return "", ""
    lines = [ln.strip() for ln in s.split("\n") if ln.strip()]
    if len(lines) >= 2:
        return lines[0], lines[1]
    one = lines[0]
    if re.search(r"[\u4e00-\u9fff]", one) and re.search(r"[A-Za-z]", one):
        for m in reversed(list(re.finditer(r"\s+(?=[A-Za-z])", one))):
            left, right = one[: m.start()].strip(), one[m.start() :].strip()
            if re.search(r"[\u4e00-\u9fff]", left) and re.search(r"[A-Za-z]", right):
                if not re.search(r"[\u4e00-\u9fff]", right):
                    return left, right
    return one, ""


def normalize_location_display(value: str) -> str:
    s = str(value or "").strip()
    # Use · only for country-location joins; keep lexical hyphens in English names intact.
    s = re.sub(r"(?<![A-Za-z])\s*-\s*(?![A-Za-z])", "·", s)
    s = re.sub(r"\s*·\s*", "·", s)
    s = re.sub(r"\s*/\s*", " / ", s)
    return re.sub(r"\s{2,}", " ", s).strip()


def en_to_location_en(en_line: str) -> str:
    if not en_line:
        return ""
    return normalize_location_display(en_line)


def en_to_country_location(en_line: str) -> tuple[str, str]:
    """Simple split for books without `sites`."""
    if not en_line.strip():
        return "", ""
    if ", " in en_line and " - " not in en_line:
        return en_line.strip(), ""
    parts = re.split(r"\s*-\s*", en_line, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return en_line.strip(), ""


def load_sheet_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "总表" not in wb.sheetnames:
        raise ValueError(f"No 总表 sheet in {xlsx_path}")
    ws = wb["总表"]
    rows = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 1).value
        author = ws.cell(r, 2).value
        year = ws.cell(r, 3).value
        publisher = ws.cell(r, 4).value
        country_loc = ws.cell(r, 5).value
        summary = ws.cell(r, 6).value
        if not (title and author and year):
            continue
        rows.append(
            {
                "title": title,
                "author": author,
                "year": int(year),
                "publisher": str(publisher or "").strip(),
                "country_loc": country_loc,
                "summary": str(summary or "").strip(),
            }
        )
    return rows


def load_books():
    content = BOOKS_PATH.read_text(encoding="utf-8")
    if not content.startswith(PREFIX):
        raise ValueError("books-data.js format changed: missing expected prefix")
    payload = content[len(PREFIX) :].strip()
    if payload.endswith(";"):
        payload = payload[:-1]
    return json.loads(payload)


def save_books(books):
    output = PREFIX + json.dumps(books, ensure_ascii=False, indent=2) + ";\n"
    BOOKS_PATH.write_text(output, encoding="utf-8")


def next_ethnography_id(books):
    best = 0
    for b in books:
        m = re.match(r"^ethnography-(\d+)$", str(b.get("id", "")))
        if m:
            best = max(best, int(m.group(1)))
    return f"ethnography-{best + 1:02d}"


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(ROOT / "民族志信息表-0513.xlsx"),
        help="Path to 民族志信息表 xlsx",
    )
    args = ap.parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        raise SystemExit(f"Missing xlsx: {xlsx_path}")

    sheet_rows = load_sheet_rows(xlsx_path)
    books = load_books()

    by_key = {}
    for b in books:
        k = match_key_from_book(b)
        by_key.setdefault(k, []).append(b)

    updated = 0
    missing = []

    for row in sheet_rows:
        k = match_key_from_row(row["title"], row["author"], row["year"])
        zh_loc, en_loc = parse_country_location(row["country_loc"])
        loc_en = en_to_location_en(en_loc)
        cn, en = parse_cn_en_title(row["title"])
        if not cn or not en:
            raise SystemExit(f"Bad title format (need 《cn》(en)): {row['title']!r}")
        title_canon = canonical_title(cn, en)

        candidates = by_key.get(k)
        if not candidates:
            missing.append(row)
            continue
        book = candidates[0]
        book["title"] = title_canon
        book["author"] = str(row["author"]).strip()
        book["year"] = row["year"]
        book["publisher"] = row["publisher"] or book.get("publisher", "")
        book["summary"] = row["summary"]
        book["sourceField"] = zh_loc or book.get("sourceField", "")
        book["locationEn"] = loc_en or book.get("locationEn", "")
        if not book.get("sites"):
            cor, loc = en_to_country_location(en_loc)
            if cor:
                book["countryOrRegion"] = cor
            if loc or not cor:
                book["location"] = loc if loc else cor
        updated += 1

    for row in missing:
        cn, en = parse_cn_en_title(row["title"])
        title_canon = canonical_title(cn, en)
        zh_loc, en_loc = parse_country_location(row["country_loc"])
        loc_en = en_to_location_en(en_loc)
        cor, loc = en_to_country_location(en_loc)
        new_id = next_ethnography_id(books)
        if en == "The Land of Open Graves":
            entry = {
                "id": new_id,
                "title": title_canon,
                "year": row["year"],
                "author": str(row["author"]).strip(),
                "publisher": row["publisher"],
                "summary": row["summary"],
                "location": "Arizona / Sonora",
                "countryOrRegion": "United States / Mexico",
                "lat": 31.3322,
                "lon": -110.9378,
                "sourceField": zh_loc,
                "locationEn": loc_en,
                "sites": [
                    {"lat": 32.2226, "lon": -110.9747},
                    {"lat": 29.0892, "lon": -110.9613},
                ],
            }
        elif en == "Global Body Shopping":
            entry = {
                "id": new_id,
                "title": title_canon,
                "year": row["year"],
                "author": str(row["author"]).strip(),
                "publisher": row["publisher"],
                "summary": row["summary"],
                "location": "Hyderabad / Australia",
                "countryOrRegion": "India / Australia",
                "lat": 17.385,
                "lon": 78.4867,
                "sourceField": zh_loc,
                "locationEn": loc_en,
                "sites": [
                    {"lat": 17.385, "lon": 78.4867},
                    {"lat": -33.8688, "lon": 151.2093},
                ],
            }
        else:
            raise SystemExit(f"No book match and no template for: {en!r}")

        books.append(entry)
        print(f"appended new book {new_id}: {title_canon}")

    save_books(books)
    print(f"xlsx_rows: {len(sheet_rows)}")
    print(f"updated_existing: {updated}")
    print(f"appended: {len(missing)}")


if __name__ == "__main__":
    main()
