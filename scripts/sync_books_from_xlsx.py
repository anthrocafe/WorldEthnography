#!/usr/bin/env python3
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "民族志信息表.xlsx"
BOOKS_PATH = ROOT / "books-data.js"
PREFIX = "window.ETHNOGRAPHY_BOOKS = "


def normalize_title(raw_title):
    return re.sub(r"\s+", " ", str(raw_title or "").replace("（", "(").replace("）", ")").strip())


def parse_cn_en_title(raw_title):
    text = normalize_title(raw_title)
    match = re.match(r"^\s*《\s*(.*?)\s*》\s*\((.*?)\)\s*$", text)
    if not match:
        return None, None
    return match.group(1).strip(), match.group(2).strip()


def row_key(title, author, year):
    return (normalize_title(title), str(author or "").strip(), int(year))


def load_sheet_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    rows = []
    cn_by_en = {}
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(2, ws.max_row + 1):
            title = ws.cell(row_idx, 1).value
            author = ws.cell(row_idx, 2).value
            year = ws.cell(row_idx, 3).value

            if not (title and author and year):
                continue

            cn, en = parse_cn_en_title(title)
            title_str = f"《{cn}》({en})" if cn and en else normalize_title(title)
            author_str = str(author).strip()
            year_int = int(year)
            key = row_key(title_str, author_str, year_int)
            if key in seen:
                continue
            seen.add(key)
            rows.append((title_str, author_str, year_int))

            if cn and en:
                cn_by_en[en] = cn

    return rows, cn_by_en


def load_books():
    content = BOOKS_PATH.read_text(encoding="utf-8")
    if not content.startswith(PREFIX):
        raise ValueError("books-data.js format changed: missing expected prefix")
    payload = content[len(PREFIX) :].strip()
    if payload.endswith(";"):
        payload = payload[:-1]
    return json.loads(payload)


def save_books(books):
    output = PREFIX + json.dumps(books, ensure_ascii=False, indent=2) + ";\n"
    BOOKS_PATH.write_text(output, encoding="utf-8")


def main():
    sheet_rows, cn_by_en = load_sheet_rows()
    keep_counter = Counter(sheet_rows)
    books = load_books()

    removed = []
    kept = []
    used_counter = Counter()

    for book in books:
        key = row_key(book.get("title", ""), book.get("author", ""), book.get("year", 0))
        if used_counter[key] < keep_counter[key]:
            kept.append(book)
            used_counter[key] += 1
        else:
            removed.append(
                (book.get("id"), book.get("title"), book.get("author"), book.get("year"))
            )

    title_updates = []
    title_pattern = re.compile(r"^《([^》]+)》\(([^)]+)\)$")
    for book in kept:
        title = str(book.get("title", "")).strip()
        m = title_pattern.match(title)
        if not m:
            continue
        old_cn, en = m.group(1).strip(), m.group(2).strip()
        new_cn = cn_by_en.get(en)
        if new_cn and new_cn != old_cn:
            book["title"] = f"《{new_cn}》({en})"
            title_updates.append((en, old_cn, new_cn))

    save_books(kept)

    final_rows = [
        row_key(b.get("title", ""), b.get("author", ""), b.get("year", 0))
        for b in kept
    ]
    is_exact = Counter(final_rows) == keep_counter

    print(f"xlsx_rows: {len(sheet_rows)}")
    print(f"books_before: {len(books)}")
    print(f"books_after: {len(kept)}")
    print(f"removed_count: {len(removed)}")
    for item in removed:
        print(f"  removed: {item}")
    print(f"title_updates: {len(title_updates)}")
    for en, old_cn, new_cn in title_updates:
        print(f"  {en}: {old_cn} -> {new_cn}")
    print(f"exact_match_after_sync: {is_exact}")


if __name__ == "__main__":
    main()

